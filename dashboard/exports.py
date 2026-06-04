"""Export builders — CSV / Excel / PDF from DB data.

All builders return bytes ready to stream as a response. No file I/O.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import desc, func as sa_func, select

from db.models import Incident, Run, UptimeCheck, get_session, init_db
from monitoring.incident_alerts import hub as alert_hub
from monitoring.uptime_persist import uptime_percent
from utils.text import strip_diacritics as _ascii
from utils.time_ro import format_ro, humanize_duration, to_ro


UPTIME_CATEGORIES = ("site_down", "site_slow", "page_down", "page_slow")
UPTIME_KIND_LABELS_RO = {
    "site_down": "Site indisponibil",
    "site_slow": "Site lent",
    "page_down": "Pagina indisponibila",
    "page_slow": "Pagina lenta",
}


# CWE-1236 — cells starting with any of these are interpreted as a formula by
# Excel / LibreOffice / Google Sheets when the file is opened. Data scraped from
# the monitored site (page titles, summaries) is untrusted, so we neutralize it.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _formula_safe(value) -> str:
    """Neutralize spreadsheet formula injection (CWE-1236).

    If the cell text begins with a formula trigger, prefix a single quote so the
    spreadsheet treats it as literal text. Applied to EVERY externally-sourced
    cell in both CSV and XLSX exports.
    """
    if value is None:
        return ""
    s = str(value)
    if s and s[0] in _FORMULA_TRIGGERS:
        return "'" + s
    return s


def _h(s) -> str:
    """ASCII-only + formula-safe helper for PDF/Excel cells.

    Strips diacritics (Helvetica/reportlab can't render ăâîșț) AND neutralizes
    formula injection. Every XLSX string cell routes through this.
    """
    return _formula_safe(_ascii(s))


# ---------------------------------------------------------------------
# data fetching helpers
# ---------------------------------------------------------------------
def _fetch_incidents(target_url: str | None, status: str = "all") -> list[Incident]:
    init_db()
    with get_session() as session:
        stmt = select(Incident)
        if target_url:
            stmt = stmt.where(Incident.target_url == target_url.rstrip("/"))
        if status == "open":
            stmt = stmt.where(Incident.resolved_at.is_(None))
        elif status == "resolved":
            stmt = stmt.where(Incident.resolved_at.is_not(None))
        stmt = stmt.order_by(desc(Incident.last_seen_at))
        return session.scalars(stmt).all()


def _fetch_runs(target_url: str | None) -> list[Run]:
    init_db()
    with get_session() as session:
        stmt = select(Run)
        if target_url:
            stmt = stmt.where(Run.base_url == target_url.rstrip("/"))
        stmt = stmt.order_by(desc(Run.started_at))
        return session.scalars(stmt).all()


def _fetch_uptime(target_url: str | None, limit: int = 10000) -> list[UptimeCheck]:
    init_db()
    with get_session() as session:
        stmt = select(UptimeCheck)
        if target_url:
            stmt = stmt.where(UptimeCheck.target_url == target_url.rstrip("/"))
        stmt = stmt.order_by(desc(UptimeCheck.checked_at)).limit(limit)
        return session.scalars(stmt).all()


def _fmt_dt(dt: datetime | None) -> str:
    """Romania-time formatting for exports — auditors read it as local time."""
    if not dt:
        return ""
    return format_ro(dt)


# ---------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------
def build_incidents_csv(target_url: str | None, status: str = "all") -> bytes:
    incidents = _fetch_incidents(target_url, status)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        "id", "target_url", "category", "severity", "page_url",
        "summary", "first_seen_at", "last_seen_at", "resolved_at",
        "is_open", "fingerprint",
    ])
    for inc in incidents:
        # _formula_safe on every externally-sourced field (CWE-1236).
        w.writerow([
            inc.id,
            _formula_safe(inc.target_url),
            _formula_safe(inc.category),
            _formula_safe(inc.severity),
            _formula_safe(inc.page_url or ""),
            _formula_safe(inc.summary),
            _fmt_dt(inc.first_seen_at),
            _fmt_dt(inc.last_seen_at),
            _fmt_dt(inc.resolved_at),
            "yes" if inc.is_open else "no",
            _formula_safe(inc.fingerprint),
        ])
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


# ---------------------------------------------------------------------
# Excel (multi-sheet)
# ---------------------------------------------------------------------
def _style_header_row(ws, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="1E3A8A")
    font = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="left", vertical="center")
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
        c.alignment = align
    ws.row_dimensions[1].height = 22


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def build_incidents_xlsx(target_url: str | None) -> bytes:
    incidents = _fetch_incidents(target_url, "all")
    runs = _fetch_runs(target_url)
    alerts = alert_hub.recent(limit=500)

    wb = Workbook()

    # =================================================================
    # Sheet 1: Indisponibilități (the auditor's first stop)
    # Just site_down / site_slow / page_down / page_slow with computed duration.
    # =================================================================
    ws_down = wb.active
    ws_down.title = "Indisponibilitati"
    ws_down.append([
        _h("Tip"), _h("URL"), _h("Inceput"), _h("Sfarsit"), _h("Durata"),
        _h("Status (la cadere)"), _h("Detalii"),
    ])
    _style_header_row(ws_down, 7)
    now = datetime.now(timezone.utc)
    downtime_rows = sorted(
        [i for i in incidents if i.category in UPTIME_CATEGORIES],
        key=lambda i: (i.first_seen_at or now),
        reverse=True,
    )
    for inc in downtime_rows:
        end = inc.resolved_at or now
        start = inc.first_seen_at or end
        if start and start.tzinfo is None:
            start = start.replace(tzinfo=timezone.utc)
        if end and end.tzinfo is None:
            end = end.replace(tzinfo=timezone.utc)
        dur_s = (end - start).total_seconds() if start else 0
        status_str = "-"
        details = inc.details or {}
        if details.get("status_code") is not None:
            status_str = f"HTTP {details['status_code']}"
        elif details.get("error"):
            status_str = (details["error"] or "")[:60]
        ws_down.append([
            _h(UPTIME_KIND_LABELS_RO.get(inc.category, inc.category)),
            _h(inc.page_url or inc.target_url),
            _h(_fmt_dt(inc.first_seen_at)),
            _h(_fmt_dt(inc.resolved_at) if inc.resolved_at else "(inca deschis)"),
            _h(humanize_duration(dur_s)),
            _h(status_str),
            _h((inc.summary or "")[:200]),
        ])
    ws_down.freeze_panes = "A2"
    _autosize(ws_down)

    # =================================================================
    # Sheet 2: Incidents (all, including non-uptime ones)
    # =================================================================
    ws = wb.create_sheet("Incidents")
    ws.append([
        "ID", "Target", "Category", "Severity", "Page", "Summary",
        "First seen", "Last seen", "Resolved at", "Open?",
        "Days open", "Fingerprint",
    ])
    _style_header_row(ws, 12)
    now = datetime.now(timezone.utc)
    for inc in incidents:
        first = inc.first_seen_at
        end = inc.resolved_at or now
        if first and first.tzinfo is None:
            first = first.replace(tzinfo=timezone.utc)
        if end and end.tzinfo is None:
            end = end.replace(tzinfo=timezone.utc)
        days = round((end - first).total_seconds() / 86400, 2) if first else 0
        ws.append([
            inc.id,
            _h(inc.target_url),
            _h(inc.category),
            _h(inc.severity),
            _h(inc.page_url or ""),
            _h(inc.summary),
            _h(_fmt_dt(inc.first_seen_at)),
            _h(_fmt_dt(inc.last_seen_at)),
            _h(_fmt_dt(inc.resolved_at)),
            "yes" if inc.is_open else "no",
            days,
            _h(inc.fingerprint),
        ])
    ws.freeze_panes = "A2"
    _autosize(ws)

    # Sheet 2: Runs
    ws2 = wb.create_sheet("Runs")
    ws2.append([
        "ID", "Target", "Started at", "Finished at", "Duration (s)",
        "Status", "Health score",
        "Pages", "Broken links", "Slow pages", "Asset failures",
        "Form issues", "Doc leaks", "SEO issues", "JS errors",
    ])
    _style_header_row(ws2, 15)
    for r in runs:
        s = r.summary or {}
        ws2.append([
            r.id, _h(r.base_url), _h(_fmt_dt(r.started_at)), _h(_fmt_dt(r.finished_at)),
            r.duration_s, _h(r.status), r.health_score,
            s.get("pages"), s.get("broken_links"), s.get("slow_pages"),
            s.get("asset_failures"), s.get("form_issues"), s.get("doc_leaks"),
            s.get("seo_issues"), s.get("console_js_errors"),
        ])
    ws2.freeze_panes = "A2"
    _autosize(ws2)

    # Sheet 3: Uptime SLA
    ws_sla = wb.create_sheet("Uptime SLA")
    ws_sla.append(["Target", "Uptime 24h", "Uptime 7d", "Uptime 30d", "Total probes (30d)"])
    _style_header_row(ws_sla, 5)
    targets_seen: list[str] = []
    if target_url:
        targets_seen = [target_url.rstrip("/")]
    else:
        # build target list from probes
        with get_session() as session:
            targets_seen = sorted({t for t, in session.execute(
                select(UptimeCheck.target_url).distinct()
            ).all()})
    for t in targets_seen:
        s24 = uptime_percent(t, hours=24)
        s7 = uptime_percent(t, hours=24 * 7)
        s30 = uptime_percent(t, hours=24 * 30)
        with get_session() as session:
            total = session.scalar(
                select(sa_func.count(UptimeCheck.id))
                .where(UptimeCheck.target_url == t)
            ) or 0
        ws_sla.append([
            _h(t),
            f"{s24}%" if s24 is not None else "-",
            f"{s7}%" if s7 is not None else "-",
            f"{s30}%" if s30 is not None else "-",
            total,
        ])
    ws_sla.freeze_panes = "A2"
    _autosize(ws_sla)

    # Sheet 4: Uptime probes (raw)
    ws_up = wb.create_sheet("Uptime probes")
    ws_up.append(["Target", "Checked at", "State", "HTTP", "Response (ms)", "Error"])
    _style_header_row(ws_up, 6)
    for u in _fetch_uptime(target_url, limit=10000):
        ws_up.append([
            _h(u.target_url),
            _h(_fmt_dt(u.checked_at)),
            _h(u.state),
            u.status_code,
            u.response_ms,
            _h((u.error or "")[:200]),
        ])
    ws_up.freeze_panes = "A2"
    _autosize(ws_up)

    # Sheet 5: Alerts audit log
    ws3 = wb.create_sheet("Alerts")
    ws3.append([
        "Fired at", "Kind", "Category", "Severity",
        "Incident ID", "Run ID", "Target", "Page",
        "Summary", "Dispatched to",
    ])
    _style_header_row(ws3, 10)
    for a in alerts:
        ws3.append([
            _h(a.get("fired_at", "")),
            _h(a.get("kind", "")),
            _h(a.get("category", "")),
            _h(a.get("severity", "")),
            a.get("incident_id"),
            a.get("run_id"),
            _h(a.get("target_url", "")),
            _h(a.get("page_url") or ""),
            _h(a.get("summary", "")),
            _h(", ".join(a.get("dispatched_to") or [])),
        ])
    ws3.freeze_panes = "A2"
    _autosize(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------
# PDF (executive audit report)
# ---------------------------------------------------------------------
def build_audit_pdf(target_url: str | None) -> bytes:
    incidents = _fetch_incidents(target_url, "all")
    runs = _fetch_runs(target_url)
    open_incidents = [i for i in incidents if i.is_open]
    resolved_incidents = [i for i in incidents if not i.is_open]
    latest_run = runs[0] if runs else None

    by_sev: dict[str, int] = {}
    by_cat: dict[str, int] = {}
    for inc in open_incidents:
        by_sev[inc.severity] = by_sev.get(inc.severity, 0) + 1
        by_cat[inc.category] = by_cat.get(inc.category, 0) + 1

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title=_h("BT Monitor - Raport audit"),
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=20, spaceAfter=4, textColor=colors.HexColor("#1E3A8A"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=14, spaceAfter=8, spaceBefore=18, textColor=colors.HexColor("#1E2D6B"))
    body = ParagraphStyle("body", parent=styles["BodyText"], fontSize=10, leading=14)
    meta = ParagraphStyle("meta", parent=styles["BodyText"], fontSize=9, textColor=colors.HexColor("#64748B"))

    story: list = []
    story.append(Paragraph(_h("BT Monitor - Raport audit"), h1))
    story.append(Paragraph(
        _h(f"Tinta: <b>{target_url or '(toate)'}</b> - "
           f"Generat: {_fmt_dt(datetime.now(timezone.utc))}"),
        meta,
    ))

    # --- Executive summary ---
    story.append(Paragraph(_h("Rezumat executiv"), h2))
    summary_rows = [
        [_h("Total scan-uri"), str(len(runs))],
        [_h("Incidente deschise"), str(len(open_incidents))],
        [_h("Incidente rezolvate"), str(len(resolved_incidents))],
        [_h("  - critice deschise"), str(by_sev.get("critical", 0))],
        [_h("  - serioase deschise"), str(by_sev.get("serious", 0))],
        [_h("  - moderate deschise"), str(by_sev.get("moderate", 0))],
        [_h("  - minore deschise"), str(by_sev.get("minor", 0))],
    ]
    if latest_run:
        summary_rows.extend([
            [_h("Health score ultim scan"), str(latest_run.health_score if latest_run.health_score is not None else "-")],
            [_h("Status ultim scan"), _h(str(latest_run.status))],
            [_h("Data ultim scan"), _h(_fmt_dt(latest_run.started_at))],
        ])
    t = Table(summary_rows, colWidths=[7 * cm, 8 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (0, -1), colors.HexColor("#F8FAFC")),
        ("FONTNAME",    (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("LINEBELOW",  (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E9F0")),
    ]))
    story.append(t)

    # --- Open by category ---
    if by_cat:
        story.append(Paragraph(_h("Incidente deschise pe categorie"), h2))
        cat_rows = [[_h("Categorie"), _h("Deschise")]]
        for cat, cnt in sorted(by_cat.items(), key=lambda kv: -kv[1]):
            cat_rows.append([_h(cat), str(cnt)])
        ct = Table(cat_rows, colWidths=[10 * cm, 5 * cm])
        ct.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
            ("TOPPADDING",   (0, 0), (-1, -1), 6),
            ("LINEBELOW",    (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E9F0")),
        ]))
        story.append(ct)

    # --- Indisponibilități (uptime outages) — dedicated section ---
    uptime_incidents = [i for i in incidents if i.category in UPTIME_CATEGORIES]
    if uptime_incidents:
        site_outages = [i for i in uptime_incidents if i.category in ("site_down", "site_slow")]
        page_outages = [i for i in uptime_incidents if i.category in ("page_down", "page_slow")]

        story.append(Paragraph(_h(f"Indisponibilitati site ({len(site_outages)})"), h2))
        if site_outages:
            srows = [[_h("Tip"), _h("Inceput"), _h("Sfarsit"), _h("Durata")]]
            now_utc = datetime.now(timezone.utc)
            for inc in site_outages[:30]:
                end = inc.resolved_at or now_utc
                start = inc.first_seen_at or end
                if start and start.tzinfo is None:
                    start = start.replace(tzinfo=timezone.utc)
                if end and end.tzinfo is None:
                    end = end.replace(tzinfo=timezone.utc)
                dur = humanize_duration((end - start).total_seconds() if start else 0)
                srows.append([
                    _h(UPTIME_KIND_LABELS_RO.get(inc.category, inc.category)),
                    _h(_fmt_dt(inc.first_seen_at)),
                    _h(_fmt_dt(inc.resolved_at) if inc.resolved_at else "(deschis)"),
                    _h(dur),
                ])
            t_sites = Table(srows, colWidths=[3.5 * cm, 4.5 * cm, 4.5 * cm, 2.5 * cm], repeatRows=1)
            t_sites.setStyle(TableStyle([
                ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#DC2626")),
                ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
                ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",     (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
                ("TOPPADDING",   (0, 0), (-1, -1), 5),
                ("LINEBELOW",    (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E9F0")),
            ]))
            story.append(t_sites)
        else:
            story.append(Paragraph(_h("(niciuna)"), body))

        story.append(Paragraph(_h(f"Indisponibilitati pagini din sitemap ({len(page_outages)})"), h2))
        if page_outages:
            prows = [[_h("Tip"), _h("Pagina"), _h("Inceput"), _h("Durata")]]
            now_utc = datetime.now(timezone.utc)
            for inc in page_outages[:50]:
                end = inc.resolved_at or now_utc
                start = inc.first_seen_at or end
                if start and start.tzinfo is None:
                    start = start.replace(tzinfo=timezone.utc)
                if end and end.tzinfo is None:
                    end = end.replace(tzinfo=timezone.utc)
                dur = humanize_duration((end - start).total_seconds() if start else 0)
                url_short = (inc.page_url or inc.target_url or "")
                if len(url_short) > 55:
                    url_short = url_short[:52] + "..."
                prows.append([
                    _h(UPTIME_KIND_LABELS_RO.get(inc.category, inc.category))[:18],
                    _h(url_short),
                    _h(_fmt_dt(inc.first_seen_at)),
                    _h(dur),
                ])
            if len(page_outages) > 50:
                prows.append(["...", _h(f"+{len(page_outages)-50} altele (vezi Excel)"), "", ""])
            t_pages = Table(prows, colWidths=[3 * cm, 6.5 * cm, 4 * cm, 2 * cm], repeatRows=1)
            t_pages.setStyle(TableStyle([
                ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
                ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
                ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",     (0, 0), (-1, -1), 8.5),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
                ("TOPPADDING",   (0, 0), (-1, -1), 4),
                ("LINEBELOW",    (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E9F0")),
            ]))
            story.append(t_pages)
        else:
            story.append(Paragraph(_h("(niciuna)"), body))

    # --- Critical incidents detail (non-uptime ones) ---
    crit = [i for i in open_incidents
            if i.severity == "critical" and i.category not in UPTIME_CATEGORIES]
    if crit:
        story.append(Paragraph(_h(f"Alte incidente CRITICE deschise ({len(crit)})"), h2))
        for inc in crit:
            story.append(Paragraph(
                _h(f"<b>#{inc.id} - {inc.category}</b><br/>{inc.summary}<br/>")
                + f"<font color='#64748B' size='8'>"
                + _h(f"page: {inc.page_url or '-'}<br/>"
                     f"prima oara: {_fmt_dt(inc.first_seen_at)} - ultima oara: {_fmt_dt(inc.last_seen_at)}")
                + "</font>",
                body,
            ))
            story.append(Spacer(1, 6))

    # --- Serious incidents (compact list) ---
    serious = [i for i in open_incidents if i.severity == "serious"]
    if serious:
        story.append(Paragraph(_h(f"Incidente serioase deschise ({len(serious)})"), h2))
        rows = [[_h("Categorie"), _h("Detalii")]]
        for inc in serious[:50]:
            rows.append([_h(inc.category), _h(inc.summary[:120])])
        if len(serious) > 50:
            rows.append(["...", _h(f"({len(serious) - 50} altele, vezi Excel)")])
        st = Table(rows, colWidths=[4 * cm, 11 * cm], repeatRows=1)
        st.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
            ("TOPPADDING",   (0, 0), (-1, -1), 5),
            ("VALIGN",       (0, 0), (-1, -1), "TOP"),
            ("LINEBELOW",    (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E9F0")),
        ]))
        story.append(st)

    # --- Recent runs ---
    if runs:
        story.append(PageBreak())
        story.append(Paragraph(_h("Istoric scan-uri (ultimele 30)"), h2))
        rrows = [["#", _h("Inceput"), _h("Status"), _h("Score"), _h("Durata")]]
        for r in runs[:30]:
            rrows.append([
                f"#{r.id}",
                _h(_fmt_dt(r.started_at)[:16] if r.started_at else "-"),
                _h(r.status or "-"),
                str(r.health_score) if r.health_score is not None else "-",
                f"{r.duration_s}s" if r.duration_s is not None else "-",
            ])
        rt = Table(rrows, colWidths=[2 * cm, 5 * cm, 3 * cm, 2.5 * cm, 2.5 * cm], repeatRows=1)
        rt.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
            ("TOPPADDING",   (0, 0), (-1, -1), 5),
            ("LINEBELOW",    (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E9F0")),
        ]))
        story.append(rt)

    story.append(Spacer(1, 18))
    story.append(Paragraph(
        _h("Document generat automat de <b>BT Monitor</b>. "
           "Datele provin din scan-urile efectuate de instrument; "
           "vezi exportul Excel pentru toate detaliile."),
        meta,
    ))

    doc.build(story)
    return buf.getvalue()
